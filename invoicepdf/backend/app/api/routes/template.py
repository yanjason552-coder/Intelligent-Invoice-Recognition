import json
from datetime import datetime
from typing import Any, Optional
from uuid import UUID, uuid4

import openpyxl
from fastapi import APIRouter, Body, File, Form, HTTPException, UploadFile
from sqlmodel import select

from app.api.deps import CurrentUser, SessionDep
from app.models.models import Message
from app.models.models_invoice import Template, TemplateField, TemplateVersion

router = APIRouter(prefix="/templates", tags=["templates"])


def _normalize_bool(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in ("1", "true", "yes", "y", "是", "必填", "required")


def _normalize_str(v: Any) -> str:
    return ("" if v is None else str(v)).strip()


def _safe_template_status(status: Optional[str]) -> str:
    if not status:
        return "enabled"
    if status not in ("enabled", "disabled", "deprecated"):
        raise HTTPException(status_code=400, detail="模板状态必须是 enabled/disabled/deprecated")
    return status


def _safe_data_type(dt: Optional[str]) -> str:
    if not dt:
        return "string"
    dt = dt.strip().lower()
    allowed = {"string", "number", "date", "datetime", "boolean", "enum", "object", "array"}
    return dt if dt in allowed else "string"


@router.get("")
@router.get("/")
def list_templates(
    *,
    session: SessionDep,
    skip: int = 0,
    limit: int = 20,
    q: str | None = None,
    status: str | None = None,
    template_type: str | None = None,
    current_user: CurrentUser,
) -> Any:
    """
    获取模板列表（分页）
    前端依赖返回结构：{ data: [...], count: number, skip, limit }
    """
    statement = select(Template)
    conditions: list[Any] = []
    if q:
        # 简化：名称/描述模糊匹配（description 可能为空）
        conditions.append((Template.name.contains(q)) | (Template.description.contains(q)))
    if status:
        conditions.append(Template.status == status)
    if template_type:
        conditions.append(Template.template_type == template_type)
    if conditions:
        from sqlalchemy import and_

        statement = statement.where(and_(*conditions))

    # 总数
    from sqlalchemy import func

    count_stmt = select(func.count()).select_from(Template)
    if conditions:
        from sqlalchemy import and_

        count_stmt = count_stmt.where(and_(*conditions))
    total = session.exec(count_stmt).one()

    templates = session.exec(
        statement.order_by(Template.create_time.desc()).offset(skip).limit(limit)
    ).all()

    # 批量补齐 current_version 字符串
    version_ids = [t.current_version_id for t in templates if t.current_version_id]
    version_map: dict[UUID, TemplateVersion] = {}
    if version_ids:
        versions = session.exec(select(TemplateVersion).where(TemplateVersion.id.in_(version_ids))).all()
        version_map = {v.id: v for v in versions}

    data = []
    for t in templates:
        v = version_map.get(t.current_version_id) if t.current_version_id else None
        data.append(
            {
                "id": str(t.id),
                "name": t.name,
                "template_type": t.template_type,
                "description": t.description,
                "status": t.status,
                "schema_id": str(t.default_schema_id) if t.default_schema_id else None,
                "default_schema_id": str(t.default_schema_id) if t.default_schema_id else None,
                "current_version": v.version if v else None,
                "accuracy": t.accuracy,
                "create_time": t.create_time,
                "update_time": t.update_time,
            }
        )

    return {"data": data, "count": int(total), "skip": skip, "limit": limit}


@router.get("/{template_id}")
def get_template_detail(
    *,
    session: SessionDep,
    template_id: UUID,
    current_user: CurrentUser,
) -> Any:
    template = session.get(Template, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")

    # 选取当前版本（若为空则取最新 created_at）
    version: TemplateVersion | None = None
    if template.current_version_id:
        version = session.get(TemplateVersion, template.current_version_id)
    if not version:
        version = session.exec(
            select(TemplateVersion)
            .where(TemplateVersion.template_id == template_id)
            .order_by(TemplateVersion.created_at.desc())
        ).first()

    fields = []
    if version:
        fields = session.exec(
            select(TemplateField)
            .where(TemplateField.template_version_id == version.id)
            .order_by(TemplateField.sort_order)
        ).all()

    return {
        "id": str(template.id),
        "name": template.name,
        "template_type": template.template_type,
        "description": template.description,
        "status": template.status,
        "accuracy": template.accuracy,
        "prompt": template.prompt,
        "schema_id": str(template.default_schema_id) if template.default_schema_id else None,
        "default_schema_id": str(template.default_schema_id) if template.default_schema_id else None,
        "version": (
            {
                "id": str(version.id),
                "version": version.version,
                "status": version.status,
                "schema_snapshot": version.schema_snapshot,
                "created_at": version.created_at,
                "published_at": version.published_at,
            }
            if version
            else None
        ),
        "fields": [
            {
                "id": str(f.id),
                "field_key": f.field_key,
                "field_name": f.field_name,
                "data_type": f.data_type,
                "is_required": bool(f.is_required),
                "description": f.description,
                "example": f.example,
                "validation": f.validation,
                "normalize": f.normalize,
                "prompt_hint": f.prompt_hint,
                "confidence_threshold": f.confidence_threshold,
                "parent_field_id": str(f.parent_field_id) if f.parent_field_id else None,
                "sort_order": f.sort_order,
            }
            for f in fields
        ],
        "create_time": template.create_time,
        "update_time": template.update_time,
    }


@router.get("/versions/{version_id}/fields")
def get_template_fields_by_version(
    *,
    session: SessionDep,
    version_id: UUID,
    current_user: CurrentUser,
) -> Any:
    """
    按模板版本ID获取字段定义（用于票据查询/审核的动态字段展示）
    返回：{ template_id, version, fields: [...] }
    """
    version = session.get(TemplateVersion, version_id)
    if not version:
        raise HTTPException(status_code=404, detail="模板版本不存在")

    template = session.get(Template, version.template_id)
    fields = session.exec(
        select(TemplateField)
        .where(TemplateField.template_version_id == version_id)
        .order_by(TemplateField.sort_order)
    ).all()

    return {
        "template_id": str(version.template_id),
        "template_name": template.name if template else None,
        "template_type": template.template_type if template else None,
        "version": version.version,
        "fields": [
            {
                "id": str(f.id),
                "field_key": f.field_key,
                "field_name": f.field_name,
                "data_type": f.data_type,
                "is_required": bool(f.is_required),
                "description": f.description,
                "example": f.example,
                "prompt_hint": f.prompt_hint,
                "confidence_threshold": f.confidence_threshold,
                "sort_order": f.sort_order,
                "parent_field_id": str(f.parent_field_id) if f.parent_field_id else None,
            }
            for f in fields
        ],
        "count": len(fields),
    }


@router.post("")
@router.post("/")
def create_template(
    *,
    session: SessionDep,
    current_user: CurrentUser,
    body: dict = Body(default_factory=dict),
) -> Any:
    """
    创建模板（最小实现）
    前端依赖返回：{ data: { template_id } }
    """
    payload = body or {}
    name = _normalize_str(payload.get("name"))
    template_type = _normalize_str(payload.get("template_type")) or "其他"
    description = payload.get("description")
    status = _safe_template_status(payload.get("status"))
    prompt = payload.get("prompt")  # 提示词
    if not name:
        raise HTTPException(status_code=400, detail="模板名称不能为空")

    # 处理 schema_id 字段
    default_schema_id = None
    if "schema_id" in payload:
        schema_id = payload.get("schema_id")
        if schema_id:
            from app.models.models_invoice import OutputSchema
            schema = session.get(OutputSchema, UUID(schema_id) if isinstance(schema_id, str) else schema_id)
            if not schema:
                raise HTTPException(status_code=400, detail="Schema不存在")
            default_schema_id = schema.id
    elif "default_schema_id" in payload:
        schema_id = payload.get("default_schema_id")
        if schema_id:
            from app.models.models_invoice import OutputSchema
            schema = session.get(OutputSchema, UUID(schema_id) if isinstance(schema_id, str) else schema_id)
            if not schema:
                raise HTTPException(status_code=400, detail="Schema不存在")
            default_schema_id = schema.id

    now = datetime.now()
    template = Template(
        id=uuid4(),
        name=name,
        template_type=template_type,
        description=description,
        status=status,
        default_schema_id=default_schema_id,
        prompt=prompt,
        # 注意：current_version_id 有外键约束指向 template_version.id，
        # 不能在首次 INSERT template 时就赋值一个尚不存在的版本ID，否则会触发 FK violation。
        # 这里先插入 template（current_version_id = NULL），再插入 version，最后回写 current_version_id。
        current_version_id=None,
        creator_id=current_user.id,
        create_time=now,
        update_time=None,
    )
    session.add(template)
    session.flush()

    version = TemplateVersion(
        id=uuid4(),
        template_id=template.id,
        version="v1.0.0",
        status="draft",
        schema_snapshot=None,
        accuracy=None,
        etag=None,
        locked_by=None,
        locked_at=None,
        created_by=current_user.id,
        created_at=now,
        published_at=None,
        deprecated_at=None,
    )
    session.add(version)
    session.flush()

    template.current_version_id = version.id
    session.add(template)
    session.commit()

    return {
        "message": "模板创建成功",
        "data": {
            "template_id": str(template.id),
            "template_name": template.name,
            "version_id": str(version.id),
            "version": version.version,
        },
    }


@router.put("/{template_id}")
def update_template(
    *,
    session: SessionDep,
    template_id: UUID,
    current_user: CurrentUser,
    body: dict = Body(default_factory=dict),
) -> Any:
    template = session.get(Template, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")

    payload = body or {}
    if "name" in payload:
        name = _normalize_str(payload.get("name"))
        if not name:
            raise HTTPException(status_code=400, detail="模板名称不能为空")
        template.name = name
    if "template_type" in payload:
        template.template_type = _normalize_str(payload.get("template_type")) or "其他"
    if "description" in payload:
        template.description = payload.get("description")
    if "prompt" in payload:
        template.prompt = payload.get("prompt")
    if "status" in payload:
        template.status = _safe_template_status(payload.get("status"))
    
    # 处理 schema_id 字段（前端可能发送 schema_id，需要映射到 default_schema_id）
    if "schema_id" in payload:
        schema_id = payload.get("schema_id")
        if schema_id:
            # 验证 schema 是否存在
            from app.models.models_invoice import OutputSchema
            schema = session.get(OutputSchema, UUID(schema_id) if isinstance(schema_id, str) else schema_id)
            if not schema:
                raise HTTPException(status_code=400, detail="Schema不存在")
            template.default_schema_id = schema.id
        else:
            template.default_schema_id = None
    elif "default_schema_id" in payload:
        schema_id = payload.get("default_schema_id")
        if schema_id:
            from app.models.models_invoice import OutputSchema
            schema = session.get(OutputSchema, UUID(schema_id) if isinstance(schema_id, str) else schema_id)
            if not schema:
                raise HTTPException(status_code=400, detail="Schema不存在")
            template.default_schema_id = schema.id
        else:
            template.default_schema_id = None

    # 处理字段更新
    if "fields" in payload:
        fields_payload = payload.get("fields", [])
        if fields_payload:
            # 获取当前版本
            version = None
            if template.current_version_id:
                version = session.get(TemplateVersion, template.current_version_id)
            
            if not version:
                # 如果没有版本，创建一个新版本
                version = TemplateVersion(
                    id=uuid4(),
                    template_id=template.id,
                    version="v1.0.0",
                    status="draft",
                    schema_snapshot=None,
                    accuracy=None,
                    etag=None,
                    locked_by=None,
                    locked_at=None,
                    created_by=current_user.id,
                    created_at=datetime.now(),
                    published_at=None,
                    deprecated_at=None,
                )
                session.add(version)
                session.flush()
                template.current_version_id = version.id
            
            # 删除旧字段（保留已存在的字段，更新或新增）
            existing_field_ids = {UUID(f.get("id")) for f in fields_payload if f.get("id")}
            if existing_field_ids:
                # 只删除不在新字段列表中的字段
                old_fields = session.exec(
                    select(TemplateField)
                    .where(TemplateField.template_version_id == version.id)
                ).all()
                for old_field in old_fields:
                    if old_field.id not in existing_field_ids:
                        session.delete(old_field)
            
            # 更新或创建字段
            for i, f in enumerate(fields_payload):
                fk = _normalize_str(f.get("field_key"))
                fn = _normalize_str(f.get("field_name"))
                if not fk or not fn:
                    continue
                
                field_id = None
                if f.get("id"):
                    try:
                        field_id = UUID(f.get("id")) if isinstance(f.get("id"), str) else f.get("id")
                    except Exception:
                        pass
                
                if field_id:
                    # 更新现有字段
                    existing_field = session.get(TemplateField, field_id)
                    if existing_field and existing_field.template_version_id == version.id:
                        existing_field.field_key = fk
                        existing_field.field_name = fn
                        existing_field.data_type = _safe_data_type(_normalize_str(f.get("data_type")))
                        existing_field.is_required = _normalize_bool(f.get("is_required"))
                        existing_field.description = f.get("description")
                        existing_field.example = f.get("example")
                        existing_field.validation = f.get("validation")
                        existing_field.normalize = f.get("normalize")
                        existing_field.prompt_hint = f.get("prompt_hint")
                        existing_field.confidence_threshold = f.get("confidence_threshold")
                        sort_order = f.get("sort_order")
                        try:
                            existing_field.sort_order = int(sort_order) if sort_order is not None else i
                        except Exception:
                            existing_field.sort_order = i
                        session.add(existing_field)
                        continue
                
                # 创建新字段
                dt = _safe_data_type(_normalize_str(f.get("data_type")))
                is_req = _normalize_bool(f.get("is_required"))
                sort_order = f.get("sort_order")
                try:
                    sort_order_int = int(sort_order) if sort_order is not None else i
                except Exception:
                    sort_order_int = i
                
                new_field = TemplateField(
                    id=uuid4(),
                    template_id=template.id,
                    template_version_id=version.id,
                    field_key=fk,
                    field_name=fn,
                    data_type=dt,
                    is_required=is_req,
                    required=is_req,
                    default_value=None,
                    description=f.get("description"),
                    example=f.get("example"),
                    validation=f.get("validation"),
                    validation_rules=None,
                    normalize=f.get("normalize"),
                    prompt_hint=f.get("prompt_hint"),
                    confidence_threshold=f.get("confidence_threshold"),
                    canonical_field=None,
                    parent_field_id=None,
                    deprecated=False,
                    deprecated_at=None,
                    position=None,
                    display_order=None,
                    sort_order=sort_order_int,
                    remark=None,
                    create_time=datetime.now(),
                )
                session.add(new_field)

    template.update_time = datetime.now()
    session.add(template)
    session.commit()
    session.refresh(template)
    return Message(message="模板更新成功")


@router.delete("/{template_id}", response_model=Message)
def delete_template(
    *,
    session: SessionDep,
    template_id: UUID,
    current_user: CurrentUser,
) -> Message:
    template = session.get(Template, template_id)
    if not template:
        raise HTTPException(status_code=404, detail="模板不存在")
    session.delete(template)
    session.commit()
    return Message(message="模板已删除")


@router.post("/import")
async def import_template(
    *,
    session: SessionDep,
    current_user: CurrentUser,
    file: UploadFile = File(...),
    name: str = Form(...),
    template_type: str = Form("其他"),
    description: str | None = Form(None),
) -> Any:
    """
    导入模板：支持 Excel(.xlsx/.xls) 与 JSON(.json)
    返回：{ data: { template_id, template_name, fields_count } }
    """
    if not file:
        raise HTTPException(status_code=400, detail="缺少文件")

    filename = (file.filename or "").lower()
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="文件为空")

    fields_payload: list[dict] = []

    # 1) JSON
    if filename.endswith(".json"):
        try:
            obj = json.loads(raw.decode("utf-8"))
        except Exception:
            # 尝试 gbk
            obj = json.loads(raw.decode("gbk", errors="ignore"))
        if isinstance(obj, dict) and isinstance(obj.get("fields"), list):
            fields_payload = obj["fields"]
        elif isinstance(obj, list):
            fields_payload = obj
        else:
            raise HTTPException(status_code=400, detail="JSON 格式不支持：需要 fields 数组或数组根结构")

    # 2) Excel
    elif filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            wb = openpyxl.load_workbook(filename=file.file, data_only=True)
        except Exception:
            # openpyxl 不能直接用 UploadFile stream（之前已 read），改用 bytes
            from io import BytesIO

            wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="Excel 无数据")

        header = [(_normalize_str(c) or "").strip() for c in (rows[0] or [])]
        header_map = {h: idx for idx, h in enumerate(header) if h}

        def col(*names: str) -> int | None:
            for n in names:
                if n in header_map:
                    return header_map[n]
            return None

        idx_key = col("field_key", "字段标识", "字段Key", "key")
        idx_name = col("field_name", "字段名称", "名称", "name")
        idx_type = col("data_type", "数据类型", "类型", "type")
        idx_req = col("is_required", "必填", "required")
        idx_desc = col("description", "描述", "说明")
        idx_ex = col("example", "示例值", "示例")
        idx_hint = col("prompt_hint", "提示", "提示词", "hint")
        idx_conf = col("confidence_threshold", "置信度阈值", "threshold")
        idx_sort = col("sort_order", "排序", "order")

        for r in rows[1:]:
            if not r:
                continue
            fk = _normalize_str(r[idx_key]) if idx_key is not None and idx_key < len(r) else ""
            fn = _normalize_str(r[idx_name]) if idx_name is not None and idx_name < len(r) else ""
            if not fk or not fn:
                continue
            dt = _safe_data_type(_normalize_str(r[idx_type]) if idx_type is not None and idx_type < len(r) else None)
            is_req = _normalize_bool(r[idx_req]) if idx_req is not None and idx_req < len(r) else False
            desc = _normalize_str(r[idx_desc]) if idx_desc is not None and idx_desc < len(r) else ""
            ex = _normalize_str(r[idx_ex]) if idx_ex is not None and idx_ex < len(r) else ""
            hint = _normalize_str(r[idx_hint]) if idx_hint is not None and idx_hint < len(r) else ""
            conf = None
            if idx_conf is not None and idx_conf < len(r) and r[idx_conf] is not None:
                try:
                    conf = float(r[idx_conf])
                except Exception:
                    conf = None
            sort_order = 0
            if idx_sort is not None and idx_sort < len(r) and r[idx_sort] is not None:
                try:
                    sort_order = int(r[idx_sort])
                except Exception:
                    sort_order = 0

            fields_payload.append(
                {
                    "field_key": fk,
                    "field_name": fn,
                    "data_type": dt,
                    "is_required": is_req,
                    "description": desc or None,
                    "example": ex or None,
                    "prompt_hint": hint or None,
                    "confidence_threshold": conf,
                    "sort_order": sort_order,
                }
            )
    else:
        raise HTTPException(status_code=400, detail="仅支持导入 .xlsx/.xls 或 .json 文件")

    # 创建模板（导入默认停用，避免误用）
    now = datetime.now()
    template = Template(
        id=uuid4(),
        name=_normalize_str(name),
        template_type=_normalize_str(template_type) or "其他",
        description=description,
        status="disabled",
        current_version_id=None,
        creator_id=current_user.id,
        create_time=now,
        update_time=None,
    )
    session.add(template)
    session.flush()

    version = TemplateVersion(
        id=uuid4(),
        template_id=template.id,
        version="v1.0.0",
        status="draft",
        schema_snapshot=None,
        accuracy=None,
        etag=None,
        locked_by=None,
        locked_at=None,
        created_by=current_user.id,
        created_at=now,
        published_at=None,
        deprecated_at=None,
    )
    session.add(version)
    session.flush()
    template.current_version_id = version.id
    session.add(template)

    # 保存字段
    fields_count = 0
    for i, f in enumerate(fields_payload):
        fk = _normalize_str(f.get("field_key"))
        fn = _normalize_str(f.get("field_name"))
        if not fk or not fn:
            continue
        dt = _safe_data_type(_normalize_str(f.get("data_type")))
        is_req = _normalize_bool(f.get("is_required"))
        sort_order = f.get("sort_order")
        try:
            sort_order_int = int(sort_order) if sort_order is not None else i
        except Exception:
            sort_order_int = i

        field = TemplateField(
            id=uuid4(),
            template_id=template.id,
            template_version_id=version.id,
            field_key=fk,
            field_name=fn,
            data_type=dt,
            is_required=is_req,
            required=is_req,  # 兼容旧字段
            default_value=None,
            description=f.get("description"),
            example=f.get("example"),
            validation=f.get("validation"),
            validation_rules=None,
            normalize=f.get("normalize"),
            prompt_hint=f.get("prompt_hint"),
            confidence_threshold=f.get("confidence_threshold"),
            canonical_field=None,
            parent_field_id=None,
            deprecated=False,
            deprecated_at=None,
            position=None,
            display_order=None,
            sort_order=sort_order_int,
            remark=None,
            create_time=now,
        )
        session.add(field)
        fields_count += 1

    session.commit()

    return {
        "message": "导入成功",
        "data": {
            "template_id": str(template.id),
            "template_name": template.name,
            "fields_count": fields_count,
        },
    }


